from io import BytesIO
import logging
import re
from pathlib import Path
from typing import List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

from exceldoc_to_markdown.config import Config


logger = logging.getLogger(__name__)

CODE_PATTERN = re.compile(r"[{}();]|^\s*(if|for|while|class|def)\b")
HEADING_PATTERN = re.compile(r"^(\d+(?:-\d+)*)\.\s*(.+)")
LIST_PATTERN = re.compile(r"^[■●・]\s*(.+)")


class ExceldocToMarkdown:
    @staticmethod
    def main(config: Config) -> bool:
        excel_path: Path | None = config.excel_path
        if excel_path is None:
            logging.error("Excelファイルのパスが設定されていません。required_checkで入力を補完してください。")
            return False

        if not excel_path.exists():
            logging.error("Excelファイルが見つかりません: %s", excel_path)
            return False

        try:
            convert_excel(excel_path, config.output_dir)
        except Exception as exc:
            logging.exception("Excel変換でエラーが発生しました: %s", exc)
            return False

        return True


def normalize_sheet_title(title: str) -> str:
    return re.sub(r"[^\w\-一-龥ぁ-んァ-ン]", "_", title)


def is_code_line(line: str) -> bool:
    return bool(CODE_PATTERN.search(line))


def convert_heading(line: str) -> str | None:
    match = HEADING_PATTERN.match(line)
    if not match:
        return None

    number_part = match.group(1)
    text = match.group(2)

    level = number_part.count("-") + 1
    normalized_number = number_part.replace("-", ".")
    hashes = "#" * level

    return f"{hashes} {normalized_number}. {text}"


def convert_list(line: str) -> str | None:
    match = LIST_PATTERN.match(line)
    if not match:
        return None
    return f"- {match.group(1)}"


def extract_images(sheet: Worksheet, image_dir: Path) -> List[str]:
    paths: List[str] = []

    if not getattr(sheet, "_images", None):
        return paths

    image_dir.mkdir(parents=True, exist_ok=True)

    for index, image in enumerate(sheet._images, start=1):
        image_path = image_dir / f"image{index}.png"

        image_bytes = image._data()
        pil_image = Image.open(BytesIO(image_bytes))
        pil_image.save(image_path)

        paths.append(str(image_path))

    return paths


def _rotate_old_output(base_output_dir: Path) -> None:
    if not base_output_dir.exists():
        return

    suffix = 1
    while True:
        candidate = base_output_dir.with_name(f"{base_output_dir.name}_old{suffix}")
        if not candidate.exists():
            break
        suffix += 1

    try:
        base_output_dir.rename(candidate)
    except Exception as exc:
        logger.error(
            "既存出力ディレクトリの退避に失敗しました: %s -> %s (%s)",
            base_output_dir,
            candidate,
            exc,
        )
        raise


def convert_sheet_to_markdown(sheet: Worksheet, output_dir: Path) -> str:
    lines: List[str] = []
    lines.append(f"# {sheet.title}")
    lines.append("")

    in_code_block = False

    for row in sheet.iter_rows(values_only=True):
        texts = [str(cell).strip() for cell in row if cell not in (None, "")]
        if not texts:
            if in_code_block:
                lines.append("``````")
                in_code_block = False
            lines.append("")
            continue

        line = " ".join(texts)

        heading = convert_heading(line)
        if heading:
            if in_code_block:
                lines.append("``````")
                in_code_block = False
            lines.append(heading)
            continue

        bullet = convert_list(line)
        if bullet:
            if in_code_block:
                lines.append("``````")
                in_code_block = False
            lines.append(bullet)
            continue

        if is_code_line(line):
            if not in_code_block:
                lines.append("``````")
                in_code_block = True
            lines.append(line)
            continue

        if in_code_block:
            lines.append("``````")
            in_code_block = False

        lines.append(line)

    if in_code_block:
        lines.append("``````")

    safe_title = normalize_sheet_title(sheet.title)
    image_dir = output_dir / f"{safe_title}_images"
    image_paths = extract_images(sheet, image_dir)

    if image_paths:
        lines.append("")
        for path in image_paths:
            relative_path = Path(path).relative_to(output_dir)
            lines.append(f"![]({relative_path.as_posix()})")

    return "\n".join(lines)


def convert_excel(excel_path: Path, output_dir: Path) -> None:
    logger.info("Excelファイルを読み込みます: %s", excel_path)
    workbook = load_workbook(excel_path, data_only=True)

    base_output_dir = output_dir / excel_path.stem
    _rotate_old_output(base_output_dir)
    base_output_dir.mkdir(parents=True, exist_ok=True)

    for sheet in workbook.worksheets:
        logger.info("シート変換開始: %s", sheet.title)

        markdown = convert_sheet_to_markdown(sheet, base_output_dir)
        safe_title = normalize_sheet_title(sheet.title)
        output_path = base_output_dir / f"{safe_title}.md"

        output_path.write_text(markdown, encoding="utf-8")
        logger.info("出力完了: %s", output_path)


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("excel", type=Path, help="入力Excelファイル")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("output_md"),
        help="出力ディレクトリ",
    )

    args = parser.parse_args()
    if not logging.getLogger().handlers:
        logging.basicConfig(level=logging.INFO, format="%(levelname)s:%(message)s")
    convert_excel(args.excel, args.output)


if __name__ == "__main__":
    main()
